#!/usr/bin/env python3
"""
Dynamic PDF Generator for bank branch audit reports.
Usage: python pdf_generator.py <excel_path> <output_dir> <audit_type> <config_json>

config_json must contain:
  columnMapping:
    branchGroupBy: str   - Excel col name for branch code grouping
    branchNameCol: str   - Excel col name for branch name
    stateCol: str        - Excel col name for state
    columns: list of:
      header: str        - Column header text (\n = line break)
      excelColumn: str|null - Excel column to pull, or null = blank
      width: float       - Column width in PDF points
      dataType: "text"|"number"
      headerColor: str|null  - Per-column header override hex color
  pdfStyle:
    pageOrientation: "landscape"|"portrait"
    headerColor1: str    - Default first color group hex
    headerColor2: str    - Default second color group hex
    fontSize: float
    rowHeight: float
    headerRowHeight: float
"""
import os
import sys
import json
import re
import argparse
import openpyxl
import pandas as pd

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape, portrait
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import urllib.request

# =========================================================
# FONT LOADING
# =========================================================
FONT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")
os.makedirs(FONT_DIR, exist_ok=True)

CARLITO_URL = "https://raw.githubusercontent.com/google/fonts/main/ofl/carlito/Carlito-Regular.ttf"
ARIMO_BOLD_URL = "https://raw.githubusercontent.com/googlefonts/arimo/main/fonts/ttf/Arimo-Bold.ttf"

reg_path = os.path.join(FONT_DIR, "Carlito-Regular.ttf")
bold_path = os.path.join(FONT_DIR, "Arimo-Bold.ttf")


def download_font(url, path):
    if not os.path.exists(path):
        sys.stderr.write(f"Downloading font -> {os.path.basename(path)}...\n")
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req) as resp, open(path, "wb") as out:
                out.write(resp.read())
        except Exception as e:
            sys.stderr.write(f"Failed to download font: {e}\n")


download_font(CARLITO_URL, reg_path)
download_font(ARIMO_BOLD_URL, bold_path)

try:
    pdfmetrics.registerFont(TTFont("Carlito", reg_path))
    pdfmetrics.registerFont(TTFont("Arimo-Bold", bold_path))
    FONT_REGULAR = "Carlito"
    FONT_BOLD = "Arimo-Bold"
except Exception as e:
    sys.stderr.write(f"Warning: Could not load fonts ({e}). Falling back.\n")
    FONT_REGULAR = "Helvetica"
    FONT_BOLD = "Helvetica-Bold"


# =========================================================
# HELPERS
# =========================================================
def format_number(val):
    """Remove trailing .0 for whole numbers."""
    if pd.isna(val) or val == "" or val is None:
        return ""
    try:
        fval = float(val)
        return str(int(fval)) if fval == int(fval) else str(fval)
    except (ValueError, TypeError):
        return str(val)


def hex_to_color(hex_str):
    try:
        return colors.HexColor(hex_str)
    except Exception:
        return colors.white


# =========================================================
# READ EXCEL — dynamic columns
# =========================================================
def read_excel(excel_path, col_map):
    """
    Read Excel file using dynamic column mapping.
    col_map keys used: branchGroupBy, branchNameCol, stateCol,
    and any excelColumn values in the columns array.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Collect all required Excel column names
    needed = set()
    needed.add(col_map["branchGroupBy"])
    needed.add(col_map["branchNameCol"])
    needed.add(col_map["stateCol"])
    for col in col_map.get("columns", []):
        if col.get("excelColumn"):
            needed.add(col["excelColumn"])
    needed_lower = {n.lower() for n in needed}

    target_sheet = None
    for sname in wb.sheetnames:
        ws = wb[sname]
        headers_row = [
            str(cell.value).strip().replace("\n", "") if cell.value is not None else ""
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        headers_lower = {h.lower() for h in headers_row}
        # Need at least the grouping column
        if col_map["branchGroupBy"].lower() in headers_lower:
            target_sheet = sname
            break

    if target_sheet is None:
        raise Exception(
            f"No valid sheet found. Required column '{col_map['branchGroupBy']}' not found."
        )

    ws = wb[target_sheet]
    headers = [
        str(cell.value).strip().replace("\n", "") if cell.value is not None else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    # Identify which columns are text-preserve (branch group, prospect-like, cuid-like)
    text_preserve_cols = {col_map["branchGroupBy"], col_map["branchNameCol"], col_map["stateCol"]}
    for col in col_map.get("columns", []):
        excel_col = col.get("excelColumn")
        if excel_col and col.get("dataType") == "text":
            text_preserve_cols.add(excel_col)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_data = {}
        all_none = True
        for i, cell in enumerate(row):
            if i < len(headers) and headers[i]:
                val = cell.value
                if val is not None:
                    all_none = False
                if headers[i] in text_preserve_cols:
                    row_data[headers[i]] = str(val).strip() if val is not None else ""
                else:
                    row_data[headers[i]] = val
        if not all_none:
            rows.append(row_data)

    return headers, rows


# =========================================================
# DYNAMIC PDF GENERATOR
# =========================================================
def generate_pdf(audit_type, branch_code, branch_name, state, rows, output_path, col_map, pdf_style):
    orientation = pdf_style.get("pageOrientation", "landscape")
    page_size = landscape(A4) if orientation == "landscape" else A4

    font_size = float(pdf_style.get("fontSize", 9))
    row_height = float(pdf_style.get("rowHeight", 30.5))
    header_row_height = float(pdf_style.get("headerRowHeight", 22.5))

    color1 = hex_to_color(pdf_style.get("headerColor1", "#FFFF00"))
    color2 = hex_to_color(pdf_style.get("headerColor2", "#4985E8"))

    columns = col_map.get("columns", [])

    # Sr No column is always first (fixed width 22.2)
    SR_NO_WIDTH = 22.2
    col_widths = [SR_NO_WIDTH] + [float(c.get("width", 80)) for c in columns]

    doc = SimpleDocTemplate(
        output_path,
        pagesize=page_size,
        leftMargin=50.2,
        rightMargin=50.2,
        topMargin=48.0,
        bottomMargin=15,
        title=os.path.basename(output_path),
    )

    style_hdr = ParagraphStyle(
        "ColHdr",
        fontName=FONT_BOLD,
        fontSize=font_size,
        alignment=TA_CENTER,
        leading=10,
        spaceBefore=0,
        spaceAfter=0,
        leftIndent=0,
        rightIndent=0,
        firstLineIndent=0,
    )

    table_data = []
    num_cols = len(col_widths)  # total cols including Sr No

    # ── Header rows (Audit Type / Branch Name / Branch Code / State) ──
    # Build spanning row with audit info across available columns
    if num_cols >= 7:
        # Original layout: 0-1 span "Audit Type:", 2 = value, 4-5 span "Branch Name:", 6 = value
        row1 = ["Audit Type :", "", str(audit_type), ""] + [""] * (num_cols - 7) + ["Branch Name :", "", str(branch_name)]
        row2 = ["Branch Code :", "", str(branch_code), ""] + [""] * (num_cols - 7) + ["State :", "", str(state)]
    elif num_cols >= 4:
        mid = num_cols // 2
        row1 = ["Audit Type :", str(audit_type)] + [""] * (mid - 2) + ["Branch Name :", str(branch_name)] + [""] * (num_cols - mid - 2)
        row2 = ["Branch Code :", str(branch_code)] + [""] * (mid - 2) + ["State :", str(state)] + [""] * (num_cols - mid - 2)
    else:
        row1 = [f"Audit: {audit_type}"] + [""] * (num_cols - 1)
        row2 = [f"Branch: {branch_code} — {branch_name}"] + [""] * (num_cols - 1)

    table_data.append(row1)
    table_data.append(row2)

    # ── Column headers ──
    hdr_row = [Paragraph("Sr<br/>No", style_hdr)]
    for col in columns:
        # Replace \n with <br/> for ReportLab
        hdr_text = col["header"].replace("\n", "<br/>")
        hdr_row.append(Paragraph(hdr_text, style_hdr))
    table_data.append(hdr_row)

    # ── Data rows ──
    for idx, row in enumerate(rows, 1):
        data_row = [str(idx)]
        for col in columns:
            excel_col = col.get("excelColumn")
            if excel_col:
                val = row.get(excel_col)
                if col.get("dataType") == "number":
                    data_row.append(format_number(val))
                else:
                    data_row.append(str(val).strip() if val is not None else "")
            else:
                data_row.append("")  # blank for hand-fill
        table_data.append(data_row)

    row_heights_list = [12.2, 14.2, header_row_height] + [row_height] * (len(table_data) - 3)

    table = Table(table_data, colWidths=col_widths, rowHeights=row_heights_list, repeatRows=3)

    # ── Build style commands ──
    style_cmds = []

    # Info header spans and styles
    if num_cols >= 7:
        style_cmds += [
            ("SPAN", (0, 0), (1, 0)),
            ("SPAN", (4, 0), (5, 0)) if num_cols == 7 else ("SPAN", (num_cols - 3, 0), (num_cols - 2, 0)),
            ("SPAN", (0, 1), (1, 1)),
            ("SPAN", (4, 1), (5, 1)) if num_cols == 7 else ("SPAN", (num_cols - 3, 1), (num_cols - 2, 1)),
        ]
    elif num_cols >= 4:
        mid = num_cols // 2
        style_cmds += [
            ("SPAN", (0, 0), (mid - 1, 0)),
            ("SPAN", (mid, 0), (num_cols - 1, 0)),
            ("SPAN", (0, 1), (mid - 1, 1)),
            ("SPAN", (mid, 1), (num_cols - 1, 1)),
        ]
    else:
        style_cmds += [
            ("SPAN", (0, 0), (num_cols - 1, 0)),
            ("SPAN", (0, 1), (num_cols - 1, 1)),
        ]

    style_cmds += [
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 1), FONT_BOLD),
        ("FONTSIZE", (0, 0), (-1, 1), font_size),
        ("FONTNAME", (0, 3), (-1, -1), FONT_REGULAR),
        ("FONTSIZE", (0, 3), (-1, -1), font_size),
        # Sr No column bold in data rows
        ("FONTNAME", (0, 3), (0, -1), FONT_BOLD),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LEFTPADDING", (0, 0), (-1, -1), 1),
        ("RIGHTPADDING", (0, 0), (-1, -1), 1),
        # Outer border
        ("LINEBEFORE", (0, 0), (0, -1), 0.5, colors.black),
        ("LINEAFTER", (num_cols - 1, 0), (num_cols - 1, -1), 0.5, colors.black),
        ("LINEABOVE", (0, 0), (-1, 0), 0.5, colors.black),
        ("LINEBELOW", (0, -1), (-1, -1), 0.5, colors.black),
        # Info header grid
        ("GRID", (0, 0), (num_cols - 1, 1), 0.5, colors.black),
        # Column header + data grid
        ("GRID", (0, 2), (-1, -1), 0.5, colors.black),
    ]

    # ── Per-column header colors ──
    # Group columns by color: columns with custom headerColor use that color,
    # others alternate between color1 and color2 in groups.
    # Default: first half gets color1, second half gets color2 (matches original behavior)
    total_data_cols = len(columns)
    midpoint = total_data_cols // 2

    for ci, col in enumerate(columns):
        table_col_idx = ci + 1  # +1 for Sr No
        custom_hex = col.get("headerColor")
        if custom_hex:
            bg = hex_to_color(custom_hex)
        else:
            bg = color1 if ci < midpoint else color2
        style_cmds.append(("BACKGROUND", (table_col_idx, 2), (table_col_idx, 2), bg))

    # Sr No header gets color1
    style_cmds.append(("BACKGROUND", (0, 2), (0, 2), color1))

    table.setStyle(TableStyle(style_cmds))
    doc.build([table])


# =========================================================
# PROCESS EXCEL
# =========================================================
def process_excel(excel_path, output_dir, audit_type, col_map, pdf_style):
    audit_type = str(audit_type).strip().upper()
    os.makedirs(output_dir, exist_ok=True)

    _headers, all_rows = read_excel(excel_path, col_map)

    MAX_BRANCHES = 10_000

    groups = {}
    for row in all_rows:
        branch = str(row.get(col_map["branchGroupBy"], "UNKNOWN")).strip()
        if branch in ("None", ""):
            branch = "UNKNOWN"
        groups.setdefault(branch, []).append(row)

    if len(groups) > MAX_BRANCHES:
        raise Exception(
            f"Too many branches ({len(groups)}). Maximum allowed is {MAX_BRANCHES}."
        )

    def _safe_filename(s: str, maxlen: int = 100) -> str:
        """Strip all characters that are unsafe in filenames, collapse whitespace."""
        s = re.sub(r"[^\w\s\-.]", "_", s)
        s = re.sub(r"[\s_]+", "_", s).strip("_")
        return s[:maxlen] or "UNKNOWN"

    results = []
    for branch_code, branch_rows in sorted(groups.items()):
        try:
            branch_name = str(branch_rows[0].get(col_map["branchNameCol"], "")).strip()
            state = str(branch_rows[0].get(col_map["stateCol"], "")).strip()
            safe_name = _safe_filename(branch_name) or _safe_filename(branch_code)
            output_file = os.path.join(output_dir, f"{safe_name}_{audit_type}.pdf")

            generate_pdf(
                audit_type, branch_code, branch_name, state,
                branch_rows, output_file, col_map, pdf_style
            )

            results.append({
                "filename": os.path.basename(output_file),
                "branchCode": branch_code,
                "branchName": branch_name,
                "rowCount": len(branch_rows),
                "fileSize": os.path.getsize(output_file),
            })
        except Exception as e:
            import traceback
            sys.stderr.write(f"Error for branch {branch_code}: {e}\n{traceback.format_exc()}\n")

    return results


# =========================================================
# MAIN
# =========================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("excel_path")
    parser.add_argument("output_dir")
    parser.add_argument("audit_type")
    parser.add_argument("config_json")
    args = parser.parse_args()

    config = json.loads(args.config_json)
    col_map = config.get("columnMapping", {})
    pdf_style = config.get("pdfStyle", {})

    # Fallback for legacy configs (old fixed 6-field format)
    if "columns" not in col_map:
        col_map = {
            "branchGroupBy": col_map.get("branchCode", "CurrentBranch"),
            "branchNameCol": col_map.get("branchName", "CurrentBranchName"),
            "stateCol": col_map.get("state", "State"),
            "columns": [
                {"header": "Prospectno", "excelColumn": col_map.get("prospectNo", "Prospectno"), "width": 101.1, "dataType": "text"},
                {"header": "CUID", "excelColumn": col_map.get("cuid", "CUID"), "width": 118.5, "dataType": "text"},
                {"header": "Tare Weight\nas per Bank", "excelColumn": col_map.get("tareWeight", "Tare Weight"), "width": 60.3, "dataType": "number"},
                {"header": "Tare Weight as\nper Audit", "excelColumn": None, "width": 67.2, "dataType": "text"},
                {"header": "Purity Check - 18K and\nabove 18K or Below 18K", "excelColumn": None, "width": 125.0, "dataType": "text"},
                {"header": "Remarks", "excelColumn": None, "width": 247.3, "dataType": "text"},
            ]
        }

    try:
        results = process_excel(args.excel_path, args.output_dir, args.audit_type, col_map, pdf_style)
        print(json.dumps({"success": True, "files": results}))
    except Exception as e:
        import traceback
        sys.stderr.write(traceback.format_exc())
        print(json.dumps({"success": False, "error": str(e)}))
        sys.exit(1)
